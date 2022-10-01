import os
import sys
import json
import requests
import base64
import random
import csv
import zipfile
import logzero
import ctypes
import multiprocessing
from calendar import isleap
from datetime import datetime
from operator import itemgetter
from logzero import logger
from time import sleep
from openpyxl import load_workbook
from openpyxl import Workbook
import boto3
import xmltodict
import pandas as pd
from fpdf import FPDF


#user = "bot_aae_is_r_dev"
#passw = "Vdj&Ty3?pFq;Zy"
#user="vsethi"
#passw= "Welcome1"


#oracle_url="https://efow.fs.us2.oraclecloud.com"
#oracle_url="https://efow-test.fa.us2.oraclecloud.com"
oracle_url= "https://efow-dev1.fa.us2.oraclecloud.com"


set_id = "300000001414463"
#Autopost_id = "300000150481649"
#source_id = "300000150481632"
######## DEV DATA
Autopost_id = "300000146748162"
source_id= "300000146748158"

s3_bucket = boto3.client('s3')
lambda_call = boto3.client('lambda', region_name='us-west-2')

bucket_name= "gwre-rpa-dev"


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
zip_path = os.path.join(ROOT_DIR, "zip")
os.makedirs(zip_path, exist_ok=True)
config_path = os.path.join(ROOT_DIR, "Config")
os.makedirs(config_path, exist_ok=True)
logs_path = os.path.join(ROOT_DIR, "Logs")
os.makedirs(logs_path, exist_ok=True)
data_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(data_path, exist_ok=True)
valid_response_Posted = "<ROWSET>\n<ROW>\n<P_LEDGER_ID>300000001414016</P_LEDGER_ID>\n<P_PERIOD_NAME>Aug-22</P_PERIOD_NAME>\n</ROW>\n</ROWSET>"
valid_length = len(valid_response_Posted.splitlines())
Fail_Flag_Posted = 0
unzip_path = os.path.join(ROOT_DIR, "Unzip")
os.makedirs(unzip_path, exist_ok=True)
general_data_path = os.path.join(ROOT_DIR, "..\\generaldata")
os.makedirs(general_data_path, exist_ok=True)
logzero.loglevel(logzero.INFO)
logzero.logfile(f"{logs_path}\\log.log")
logzero.json()


Day = datetime.now().strftime("%d")
months = (["Jan", "Feb", "Mar", "Apr",
           "May", "Jun", "Jul", "Aug",
           "Sep", "Oct", "Nov", "Dec"]
          )


EMEA = ["53", "55", "45", "58", "57", "43", "46", "47", "48", "49", "51", "41"]
APAC = ["71", "72", "80", "84", "85", "82", "86"]
AMER = ["16", "65", "61", "10", "12", "14", "15", "01"]
REGIONS = ["EMEA","AMER_US","AMER_NON_US","APAC"]

#next_mon = int(Month) % 12
#mon = months[next_mon]



   
def set_user(data):
    logger.info("Setting credentials")
    try:
       user,passw= data.split("///")
       with open(f"{ROOT_DIR}\\Data\\credentials.txt","wb") as cred:
            cred.write(base64.b64encode(user.encode()))
            cred.write(b"\n")
            cred.write(base64.b64encode(passw.encode()))
            cred.close()
    except:
        logger.info("Error on set credentials")
        return "Error on set credentials // //"


def read_credentials():
    try:
        logger.info("Reading credentials")
        usr_psw = open(f"{data_path}\\credentials.txt","rb")
        user = base64.b64decode(usr_psw.readline().decode()).decode()
        passw = base64.b64decode(usr_psw.readline().decode()).decode()
        usr_psw.close()
        user = user.strip()
        passw = passw.strip()
        #user , passw="vsethi","Welcome1"
    except:
        logger.info("Error on reading credentials")
        return "Error on read credentials // "
    return user,passw



def get_ledgers_info():
    try:
        logger.info("Getting ledgers info to look for closed ones")
        url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/ledgersLOV?limit=10000"
        #print(url)
        headerss = {"Content-Type": "application/json",
                "Connection":  "Keep-Alive"
                }
        user,passw  = read_credentials()
    except:
       logger.info("Ledgers info can not be read")
       return "Error on read ledger info // "
    return requests.get(url, auth=(user, passw), headers=headerss)




'''
def get_entities():
    logger.info("Reading Ledgers Id from Oracle")
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        #print(ledgers_info)
        #print(ledgers_info.text)
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        for i in ledger_json:
            ledgers[i["Name"]] = i["LedgerId"]
    except:
        return "Error on set the entities // // "
    return  {
        "10": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "12": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "14": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "15": {"Name": "GW US PL", "Ledger":  ledgers["GW US PL"]},
        "16": {"Name": "GW CA PL", "Ledger":  ledgers["GW CA PL"]},
        "65": {"Name": "GW BR PL", "Ledger":  ledgers["GW BR PL"]},
        "41": {"Name": "GW UK PL", "Ledger":  ledgers["GW UK PL"]},
        "43": {"Name": "GW FR PL", "Ledger":  ledgers["GW FR PL"]},
        "45": {"Name": "GW DE PL", "Ledger":  ledgers["GW DE PL"]},
        "46": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "47": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "48": {"Name": "GW IE PL", "Ledger":  ledgers["GW IE PL"]},
        "49": {"Name": "GW IT PL", "Ledger":  ledgers["GW IT PL"]},
        "51": {"Name": "GW PL PL", "Ledger":  ledgers["GW PL PL"]},
        "55": {"Name": "GW CH PL", "Ledger":  ledgers["GW CH PL"]},
        "57": {"Name": "GW ES PL", "Ledger":  ledgers["GW ES PL"]},
        "71": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "72": {"Name": "GW AU PL", "Ledger":  ledgers["GW AU PL"]},
        "80": {"Name": "GW CN PL", "Ledger":  ledgers["GW CN PL"]},
        "86": {"Name": "GW JP PL", "Ledger":  ledgers["GW JP PL"]},
        "84": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "85": {"Name": "GW IN PL", "Ledger":  ledgers["GW IN PL"]},
        "82": {"Name": "GW MY PL", "Ledger":  ledgers["GW MY PL"]},
        "61": {"Name": "GW AR PL", "Ledger":  ledgers["GW AR PL"]},
        "53": {"Name": "GW AT PL", "Ledger":  ledgers["GW AT PL"]},
        "58": {"Name": "GW DK PL", "Ledger":  ledgers["GW DK PL"]}, 
    }
'''




def get_entities(ledger_type):
    logger.info("Getting entities")
    logger.info("Reading Ledgers Id from Oracle")
    if ledger_type.upper() == "PRIMARY":
        ledger_file=f"{general_data_path}\\primary.txt"
    elif ledger_type.upper() == "SUBLEDGER":
        ledger_file=f"{general_data_path}\\subledger.txt"
    elif ledger_type.upper() == "PRIANDSUB":
        ledger_file=f"{general_data_path}\\priandsub.txt"
    elif ledger_type.upper() == "PRIANDSUBR":
        ledger_file=f"{general_data_path}\\priandsubr.txt"
    elif ledger_type.upper() == "REVALUATION":
        ledger_file=f"{general_data_path}\\revaluation.txt"
    elif ledger_type.upper() == "CL":
        ledger_file=f"{general_data_path}\\cl.txt"
    with open(ledger_file,"r") as ledger_config:
        ledger_list=ledger_config.read().splitlines()
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        for i in ledger_json:
            ledgers[i["Name"]] = i["LedgerId"]        
        entities={}
        for i in ledger_list:
            #print(i)
            try:
                data = i.split(",")
                #print(data)
                entity=data[0].strip()
                ledger_name = data[1].strip()
                entities.update({ entity: {"Name": ledger_name, "Ledger":  ledgers[ledger_name]} })  
            except:
                pass
    except:
        logger.info("Error on reading entities")
        return "Error on set the entities // // "
    return  entities

'''
entities = {
                        "10": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "12": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "14": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "15": {"Name": "GW US PL", "Ledger": "300000001414016"},
                        "16": {"Name": "GW CA PL", "Ledger": "300000001414017"},
                        "65": {"Name": "GW BR PL", "Ledger": "300000001414018"},
                        "41": {"Name": "GW UK PL", "Ledger": "300000001414019"},
                        "43": {"Name": "GW FR PL", "Ledger": "300000001414020"},
                        "45": {"Name": "GW DE PL", "Ledger": "300000001414021"},
                        "46": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "47": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "48": {"Name": "GW IE PL", "Ledger": "300000001414022"},
                        "49": {"Name": "GW IT PL", "Ledger": "300000001414023"},
                        "51": {"Name": "GW PL PL", "Ledger": "300000001414024"},
                        "55": {"Name": "GW CH PL", "Ledger": "300000001414025"},
                        "57": {"Name": "GW ES PL", "Ledger": "300000001414026"},
                        "71": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "72": {"Name": "GW AU PL", "Ledger": "300000001414027"},
                        "80": {"Name": "GW CN PL", "Ledger": "300000001414029"},
                        "86": {"Name": "GW JP PL", "Ledger": "300000001414030"},
                        "84": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "85": {"Name": "GW IN PL", "Ledger": "300000016748510"},
                        "82": {"Name": "GW MY PL", "Ledger": "300000019642549"},    
                        "61": {"Name": "GW AR PL", "Ledger": "300000019833039"},
                        "53": {"Name": "GW AT PL", "Ledger": "300000017944094"},
                        "58": {"Name": "GW DK PL", "Ledger": "300000026789264"},
                        "01": {"Name": "GW AR RC", "Ledger": "300000019833049"}
                        }


'''                        '''
revaluation = {
                    "43": {"Ledger": "300000014386895", "Name": "GW France Revaluation"},
                    "86": {"Ledger": "300000014386929", "Name": "GW Japan Revaluation"},
                    "51": {"Ledger": "300000014386933", "Name": "GW Poland Revaluation"},
                    "57": {"Ledger": "300000014386939", "Name": "GW Spain Revaluation"},
                    "10": {"Ledger": "300000014387011", "Name": "GW US Revaluation"},
                    "82": {"Ledger": "300000019830939", "Name": "GW Malaysia Revaluation"},
                    "01": {"Ledger": "300000019954804", "Name": "GW Argentina Revaluation"},
                    "65": {"Ledger": "300000014386883", "Name": "GW Brazil Revaluation"},
                    "16": {"Ledger": "300000014386887", "Name": "GW Canada Revaluation"},
                    "80": {"Ledger": "300000014386891", "Name": "GW China Revaluation"},
                    "84": {"Ledger": "300000017430661", "Name": "GW India Revaluation"},
                    "71": {"Ledger": "300000014386873", "Name": "GW Australia Revaluation"},
                    "45": {"Ledger": "300000014386904", "Name": "GW Germany Revaluation"},
                    "49": {"Ledger": "300000014386922", "Name": "GW Italy Revaluation"},
                    "41": {"Ledger": "300000014386953", "Name": "GW UK Revaluation"},
                    "58": {"Ledger": "300000018201450", "Name": "GW Denmark Revaluation"},
                    "47": {"Ledger": "300000014386915", "Name": "GW Ireland Revaluation"},
                    "71": {"Ledger": "300000026878446", "Name": "GW Austria Revaluation"}
}

'''

def get_ending_days(Year):
    logger.info("Getting the number of days by month")
    try:
        if isleap(int(Year)):

            ending_days = ["31", "29", "31", "30", "31", "30", "31", "31", "30", "31", "30", "31"]
            # print(f"Leap:{ending_days}")
        else:
            ending_days = ["31", "28", "31", "30", "31", "30", "31", "31", "30", "31", "30", "31"]
            # print(f"No_Leap:{ending_days}")
    except:
        logger.info("Error on setting the days")
        return "Error on set the days"
    return ending_days


def Push_To_S3(Xlsx_File, process, subdir):
    '''
    This function push a file into the S3 bucket
    '''
    logger.info("Pushing to s3")
    try:
        with open(Xlsx_File, "rb") as f:
            key = Xlsx_File.split("\\")[-1]
            response = s3_bucket.upload_fileobj(f, bucket_name, f"{process}/{subdir}/{key}")
            f.close()
    except Exception as e:
        logger.info(f"Error occurs while uploading to s3")
        return "Python Error : Error occurs while uploading <BR>"


def Create_Validate_SOAP(ledger_id, period):
    return f'''<soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" xmlns:pub="http://xmlns.oracle.com/oxp/service/PublicReportService">
   <soap:Header/>
   <soap:Body>
      <pub:runReport>
         <pub:reportRequest>
            <pub:attributeFormat>xml</pub:attributeFormat>
            <pub:flattenXML>true</pub:flattenXML>
            <pub:parameterNameValues>
                  <pub:item>
                  <pub:name>P_LEDGER_ID</pub:name>
                  <pub:values>
                      <pub:item>{ledger_id}</pub:item>
                  </pub:values>
                  </pub:item>
                  <pub:item>
                  <pub:name>P_PERIOD_NAME</pub:name>
                  <pub:values>
                      <pub:item>{period}</pub:item>
                  </pub:values>
                  </pub:item>
            </pub:parameterNameValues>
            <pub:reportAbsolutePath>/Custom/Financials/General Ledger/GW GL Unposted Journals Report.xdo</pub:reportAbsolutePath>
            <pub:sizeOfDataChunkDownload>-1</pub:sizeOfDataChunkDownload>
         </pub:reportRequest>
      </pub:runReport>
   </soap:Body>
</soap:Envelope>
'''


def Validate_Journals(soap):
    # print(soap)
    logger.info("Validating ledgers")
    url = f"{oracle_url}/xmlpserver/services/ExternalReportWSSService?WSDL"
    headerss = {"Content-Type": "application/soap+xml",
                "Accept-Encoding": "gzip, deflate",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    resp = requests.post(url, data=soap, auth=(user, passw), headers=headerss)
    respcod = resp.text.split("Bytes>")[1].split("</ns2")[0]
    respdec = base64.b64decode(respcod)
    return respdec


def Validate_Region(region_period):
    try:
        logger.info("Validating errors")
        entities = get_entities()
        # print(region,period)
        period,region=region_period.split(",")
        region= region.upper()
        Month, Year = period.split("-")
        if region == "EMEA":
            region_ledgers = EMEA
        elif region == "APAC":
            region_ledgers = APAC
        elif region == "AMER":
            region_ledgers = AMER
        else:
            return "Error : Not Valid Region"
        # print(region_ledgers)
        with open(f"{ROOT_DIR}\\config\\config{Month}{Year}.cfg", "a+") as config:
            config.write(f"{region}\n")
        for i in region_ledgers:
            #print(i)
            ledger = entities[i]["Ledger"]
            Entity = entities[i]["Name"]
            #print(f"Validating ledger {ledger} --> {Entity}  for region {region} ")
            soap = Create_Validate_SOAP(ledger, period)
            response = Validate_Journals(soap)
            #print(response.decode(), valid_response_Posted)
            if len(response.decode().strip().splitlines()) != valid_length:
                break
                return f"Python Error : Ledger  {ledger} of the {Entity} is not posted"
        return 0
    except:
        logger.info("Error  region or period cannot be read")
        return f"Error : Period or region could not be read"


def Create_AutoPost(Autopost_id):
    logger.info("Autoposting")
    return f'''{{
    "OperationName":"submitESSJobRequest",
    "JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common/",
    "JobDefName":"AutomaticPosting",
    "ESSParameters":"{Autopost_id}",
    "ReqstId":null
    }} '''


def GL_Request(soap):
    logger.info("Requesting to GL Interface")
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    return requests.post(url, data=soap, auth=(user, passw), headers=headerss)


def Create_Multi_Period_Accounting(ledger_id, Year, Month):
    logger.info("Creating multiperiod accounting SOAP")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/subledgerAccounting/shared",
"JobDefName":"XLAFSNAPMPA",
"ESSParameters":"M,200,{ledger_id},#NULL,{Year}00{Month},F,1000,Y,Y,#NULL,N,#NULL,#NULL,#NULL,#NULL,#NULL,#NULL,#NULL",
"ReqstId":null
}}      '''


def Multi_Period_Accounting(region, Year, Month):
    logger.info("Trying to create multiperiod accounting")
    entities = get_entities("priandsub")
    zipfile_name = ""
    Fail_Flag = 0
    if region == "EMEA":
        region_ledgers = EMEA
    elif region == "APAC":
        region_ledgers = APAC
    elif region == "AMER":
        region_ledgers = AMER
    else:
        return 1
    # print(region_ledgers)
    for i in region_ledgers:
        #print(i)
        ledger = entities[i]["Ledger"]
        Entity = entities[i]["Name"]
        #print(f"Creating Multi-Period-Accounting for ledger {ledger} --> {Entity}  for region {region} ")
        soap = Create_Multi_Period_Accounting(ledger, Year, Month)
        response = GL_Request(soap)
        # print(response.text)
        PID = json.loads(response.text)["ReqstId"]
        while True:
            Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
            if Status == "SUCCEEDED":
                break
            if Status == "ERROR" or Status == "WARNING":
                Fail_Flag = 1
                zipdata = get_files(PID, "All")
                zipfile_name = f"{ROOT_DIR}\\{PID}.zip"
                with open(f"{zipfile_name}", "wb") as zip:
                    zip.write(zipdata)
                break
    return f"{Fail_Flag} {zipfile_name}"


def Get_Status(RqstId):
    logger.info(f"Get status for PID {RqstId}")
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations?finder=ESSJobStatusRF;requestId={RqstId}"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    #print(url)
    user,passw = read_credentials()
    return requests.get(url, auth=(user, passw), headers=headerss)


def get_files(id, type):
    logger.info(f"Get files from PID {id}")
    url = f'{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations?finder=ESSJobExecutionDetailsRF;requestId={id},fileType={type}'
    #print(url)
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    req = requests.get(url, auth=(user, passw), headers=headerss).text
    # print(req)
    json_data = json.loads(req)
    print(json_data)
    data = json_data["items"][0]["DocumentContent"]
    # print(data)
    return base64.b64decode(data)


def Close_period_exception(ledger_id, period, Year, Month):
    logger.info("Verifing Close period exception")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/subledgerAccounting/shared",
"JobDefName":"XLAPEXRPT",
"ESSParameters":"#NULL,#NULL,{ledger_id},#NULL,#NULL,{period},{Year}00{Month},{period},#NULL,#NULL,#NULL",
"ReqstId":null
}}'''


def Get_Close_Exception(region, period, Year, Month):
    logger.info("Getting close exception")
    entities = get_entities("priandsub")
    zipfiles = []
    Fail_Flag = 0
    if region == "EMEA":
        region_ledgers = EMEA
    elif region == "APAC":
        region_ledgers = APAC
    elif region == "AMER":
        region_ledgers = AMER
    else:
        return 1
    # print(region_ledgers)
    count = 0
    for i in region_ledgers:
        #print(i)
        ledger = entities[i]["Ledger"]
        Entity = entities[i]["Name"]
        #print(f"Close Period for ledger {ledger} --> {Entity}  for region {region} ")
        soap = Close_period_exception(ledger, period, Year, Month)
        #print(soap)
        response = GL_Request(soap)
        #print(response.text)
        PID = json.loads(response.text)["ReqstId"]
        while True:
            Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
            if Status == "SUCCEEDED":
                break
            if Status == "ERROR" or Status == "WARNING":
                Fail_Flag = 1
                break
        if count % 2 == 0:
            zipdata = get_files(PID, "out")
        else:
            zipdata = get_files(PID, "log")
        count += 1
        zipfile_name = f"{ROOT_DIR}\\{PID}.zip"
        with open(f"{zipfile_name}", "wb") as zip:
            zip.write(zipdata)
        zipfiles.append(zipfile_name)
        zipstring = str(zipfiles)
    return f"{Fail_Flag} Zipfiles {zipstring}"


'''
def Do_Revaluation(region, Year, Month):
    Fail_Flag = 0
    zipstring = ""
    current_month = months[int(Month) - 1]
    year_two_digits = str(Year[-2:])
    Period = f"{current_month}-{year_two_digits}"
    Ending_Days = get_ending_days(Year)
    Last_Day = Ending_Days[int(Month) - 1]
    if region == "EMEA":
        region_ledgers = EMEA
    elif region == "APAC":
        region_ledgers = APAC
    elif region == "AMER":
        region_ledgers = AMER
    else:
        return 1
    for i in region_ledgers:
        try:
            ledger_id = entities[i]["Ledger"]
            Entity = entities[i]["Name"]
            rev_id = revaluation[i]["Ledger"]
        except:
            pass
        else:
            soap = rev_SOAP(ledger_id, rev_id, Period, Year, Month, Last_Day)
            print(soap)
            with open(f"{ROOT_DIR}\\rev.txt", "a+") as f:
                f.write(soap)
            response = GL_Request(soap)
            print(response.text)
            PID = json.loads(response.text)["ReqstId"]
            if int(PID) != -1:
                while True:
                    Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
                    if Status == "SUCCEEDED":
                        break
                    if Status == "ERROR" or Status == "WARNING":
                        Fail_Flag = 1
                        break
                # if count % 2 == 0:
                #    zipdata = get_files(PID,"out")
                # else:
                #    zipdata = get_files(PID,"log")
                # count += 1
                zipdata = get_files(PID, "All")
                zipfile_name = f"{ROOT_DIR}\\{PID}.zip"
                with open(f"{zipfile_name}", "wb") as zip:
                    zip.write(zipdata)
                zipfiles.append(zipfile_name)
                zipstring = str(zipfiles)
            else:
                Fail_Flag = 1
    return f"{Fail_Flag} Zipfiles {zipstring}"

'''


def rev_SOAP(ledger_id, rev_id, Period, Year, Month, Last_day):
    logger.info("Creating revaluation payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"Revaluation",
"ESSParameters":"300000001414463,{ledger_id},{rev_id},{Period},{Year}-{Month}-{Last_day},{Year}-{Month}-{Last_day}",
"ReqstId":null
}}'''


def Revaluation(ledger_id, Year, Month): 
    logger.info("Trying to do revaluation")
    year_two_digits = str(Year[-2:])
    revaluation=get_rev()
    print(revaluation)
    #try:
    #    with open(f"{data_path}\\rev_{ledger_id}{Year}{Month}.cfg","r") as rev_file:
    #       return f"{zip_path}\\rev_{ledger_id}_{Year}_{Month}.zip"
    #except:
    #    pass
    Fail_Flag = 0
    entities = get_entities("priandsubr")
    rev_entities = get_entities("revaluation")
    print(ledger_id,rev_entities)
    if ledger_id not in entities.keys():
        logger.info("Ledger id not in entities")
        return f"{zip_path}\\rev_{ledger_id}_{year_two_digits}_{Month}.zip"
    if ledger_id not in rev_entities.keys():
        logger.info("Ledger id not in revaluation entities")
        return f"{zip_path}\\rev_{ledger_id}_{year_two_digits}_{Month}.zip"
    if ledger_id not in revaluation.keys():
        logger.info("Ledger id not in revaluation from api")
        return f"{zip_path}\\rev_{ledger_id}_{year_two_digits}_{Month}.zip"
    current_month = months[int(Month) - 1]
    Period = f"{current_month}-{year_two_digits}"
    Ending_Days = get_ending_days(Year)
    Last_Day = Ending_Days[int(Month) - 1]
    ledger = entities[ledger_id]["Ledger"]
    rev_id = revaluation[ledger_id]["Ledger"]
    rev_name = revaluation[ledger_id]["Name"]
    if "SL" not in rev_name:
        soap = rev_SOAP(ledger, rev_id, Period, Year, Month, Last_Day)
        response = GL_Request(soap)
        logger.info("Verifing revalue PID")
        PID = json.loads(response.text)["ReqstId"]
        if int(PID) != -1:
            with open(f"{data_path}\\PID_{ledger_id}_{current_month}_{year_two_digits}.INFO","a+") as INFO_file:
                    INFO_file.write(f"Revaluation PID :::: {PID}\n")
            while True:
                Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
                #print(Status)
                if Status == "SUCCEEDED":
                    logger.info("Revalue Succeeded")
                    break   
                if Status == "ERROR" or Status == "WARNING":
                    Fail_Flag = 1
                    logger.info("Error or Warning on Revalue")
                    break
            logger.info("getting file data")    
            zipdata = get_files(PID, "OUT")
            zipfile_name = f"{zip_path}\\revaluation_report_{ledger_id}_{current_month}_{year_two_digits}.zip"
            with open(f"{zipfile_name}", "wb") as zip:
                zip.write(zipdata)
            if Fail_Flag == 1:
                with open(f"{zip_path}\\revaluation_error_report_{ledger_id}_{current_month}_{year_two_digits}.txt", "w") as errortxt:
                    errortxt.write(f"Error or Warning on Revaluation for PID: {PID}")
                return f"{zip_path}\\revaluation_error_report_{ledger_id}_{current_month}_{year_two_digits}.txt"
            else:
                with open(f"{data_path}\\revaluation_report_{ledger_id}_{current_month}_{year_two_digits}.cfg","a+") as rev_file:
                    rev_file.write(f"{ledger_id}-{Year}/{year_two_digits}\n")
                return f"{zipfile_name}"
        else:
            with open(f"{zip_path}\\revaluation_error_report_{ledger_id}_{current_month}_{year_two_digits}.txt", "w") as errortxt:
                errortxt.write(f"Can't get valid id revaluation for {ledger} :::::: {PID}")
            return f"{zip_path}\\revaluation_error_report_{ledger_id}_{current_month}_{year_two_digits}.txt"


def Get_translation_SOAP(Ledger_Id, Period_Name):
    logger.info("Creating Translation payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"Translation",
"ESSParameters":"{Ledger_Id},300000001414463,3002,300000001414025,USD,{Period_Name},A,#NULL,#NULL,#NULL,#NULL,#NULL,Y,300000001412001,3002X300000001412001",
"ReqstId":null
}}'''








def Do_MPA(ledger_id, Year, Month):
    year_two_digits = str(Year[-2:])
    mon = months[int(Month) - 1]
    mo = months[int(Month)  - 2]
    logger.info("Trying to do multiperiod accounting")
    try:
        with open(f"{data_path}\\mpa_{ledger_id}{Year}{Month}.cfg","r") as mpa_file:
            return f"{zip_path}\\mpa_report_{ledger_id}_{mon}_{year_two_digits}.pdf"
    except:
       pass
    zipfile_name = ""  
    entities = get_entities("priandsub")
    if ledger_id not in entities.keys():
        logger.info("LEdger id not in entities")
        return f"{zip_path}\\mpa_report_{ledger_id}_{mon}_{year_two_digits}.pdf"
    ledger = entities[ledger_id]["Ledger"]
    Entity = entities[ledger_id]["Name"]
    logger.info(f"Creating Multi-Period-Accounting for ledger {ledger} --> {Entity}  ")
    soap = Create_Multi_Period_Accounting(ledger, Year, Month)
    response = GL_Request(soap)
    PID = json.loads(response.text)["ReqstId"]
    logger.info("Verifying status for multiperios accounting ")
    if PID != -1:
        with open(f"{data_path}\\PID_{ledger_id}_{mo}_{year_two_digits}.INFO","w+") as INFO_file:
                INFO_file.write(f"MPA PID :::::  {PID}\n")
        while True:
            Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
            if Status == "SUCCEEDED" or Status == "ERROR" or Status == "WARNING":
                logger.info("Multiperiod accounting Completed")
                break
        zipdata = get_files(PID, "All")
        zipfile_name = f"{zip_path}\\mpa_report_{ledger_id}_{mon}_{year_two_digits}.zip"
        with open(f"{zipfile_name}", "wb+") as zip:
            zip.write(zipdata)
        with open(f"{data_path}\\mpa_{ledger_id}{Year}{Month}.cfg","w+") as mpa_file:
                mpa_file.write(f"{ledger_id}{Year}{Month}\n")
        #######Added
        #with open(f"{data_path}\\mpa_{ledger_id}{Year}{Month}.cfg","w+") as mpa_file:
        #        mpa_file.write(f"{ledger_id}{Year}{Month}\n")
        with zipfile.ZipFile(f"{zipfile_name}", 'r') as zip_mpa:
                zip_mpa.extractall(f"{unzip_path}")
        txt_file=f"{unzip_path}\\{PID}.txt"
        with open(txt_file,"r") as text:
            pid_to_download = text.read()    
            pid_to_download = pid_to_download.split("identifier")[1].split("for")[0].strip()
        with open(f"{data_path}\\PID_{ledger_id}_{mo}_{year_two_digits}.INFO","a+") as INFO_file:
                INFO_file.write(f"MPA Child PID :::::  {PID}\n")
        while True:
            Status = json.loads(Get_Status(pid_to_download).text)["items"][0]["RequestStatus"]
            if Status == "SUCCEEDED" or Status == "ERROR" or Status == "WARNING":
                logger.info("Verifing reporting")
                break
        print(f"\n\n\nTrying to get child data from pid {pid_to_download}\n\n\n")
        child_zipdata = get_files(pid_to_download, "All")
        print(f"\n\n\nTrying to save child\n\n\n")
        child_zipfile_name = f"{zip_path}\\mpa_{pid_to_download}.zip"
        with open(f"{child_zipfile_name}", "wb+") as zip:
            zip.write(child_zipdata)
        print(f"\n\n\n child zipdata saved as {child_zipfile_name}")
        mpa_month = months[int(Month) - 1]
        child_filename=get_mpa_report(child_zipfile_name,f"{zip_path}\\mpa_report_{ledger_id}_{mpa_month}_{year_two_digits}.pdf")
        #############
        return f"{child_filename}"
    else:
        with open(f"{data_path}\\error_on_MPA.txt","a+") as INFO_file:
                INFO_file.write(f"Error on set PID\n")
        return f"{data_path}\\error_on_MPA.txt"


def DO_CPE(ledger_id, period, Year, Month):
    year_two_digits = str(Year[-2:])
    mon = months[int(Month) - 1]
    logger.info("Doing a close period exception")
    Fail_Flag = 0
    entities = get_entities("priandsub")
    if ledger_id not in entities.keys():
        return f"{zip_path}\\CPE_report_{ledger_id}_{mon}_{year_two_digits}.zip"
    ledger = entities[ledger_id]["Ledger"]
    Entity = entities[ledger_id]["Name"]
    #print(f"Close Period exception for ledger {ledger} --> {Entity}")
    soap = Close_period_exception(ledger, period, Year, Month)
    #print(soap)
    response = GL_Request(soap)
    #print(response.text)
    PID = json.loads(response.text)["ReqstId"]
    logger.info("Verifying the status of the close period exception")
    if PID != -1:
        with open(f"{data_path}\\PID_{ledger_id}_{mon}_{year_two_digits}.INFO","a+") as INFO_file:
                INFO_file.write(f"CPE PID :::::  {PID}\n")
        while True:
            Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
            if Status == "SUCCEEDED" or Status == "ERROR" or Status == "WARNING":
                logger.info("Close perios exception completed")
                break
        zipdata = get_files(PID, "All")
        zipfile_name = f"{zip_path}\\CPE_report_{ledger_id}_{mon}_{year_two_digits}.zip"
        with open(f"{zipfile_name}", "wb") as zip:
            zip.write(zipdata)
        return f"{zipfile_name}"
    else:
        with open(f"{data_path}\\error_cpe.cfg","w") as INFO_file:
                INFO_file.write(f"Error on get PID\n")
        return f"{data_path}\\error_cpe.cfg"


def Do_Translation(ledger_id, Period):
    logger.info("Doing translation")
    Year = Period.split("-")[1]
    Month = Period.split("-")[0]
    mon = Month
    try:
        with open(f"{data_path}\\tra_{ledger_id}{Year}{Month}.cfg","r") as tra_file:
            return f"{zip_path}\\translation_report_{ledger_id}_{Month}_{Year}.zip"
    except:
        pass
    Fail_Flag = 0
    try:
        entities = get_entities("priandsub")
        ledger = entities[ledger_id]["Ledger"]
    except:
        pass
    else:
        soap = Get_translation_SOAP(ledger, Period)
        response = GL_Request(soap)
        PID = json.loads(response.text)["ReqstId"]
        logger.info("Verifying status for translation")
        if int(PID) != -1:
            with open(f"{data_path}\\PID_{ledger_id}_{mon}_{Year}.INFO","a+") as INFO_file:
                INFO_file.write(f"Translation PID :::::  {PID}\n")
            while True:
                Status = json.loads(Get_Status(PID).text)["items"][0]["RequestStatus"]
                if Status == "SUCCEEDED" or Status == "ERROR" or Status == "WARNING":
                    logger.info("Translation completed")
                    break
            zipdata = get_files(PID, "All")
            zipfile_name = f"{zip_path}\\translation_report_{ledger_id}_{mon}_{Year}.zip"
            with open(f"{zipfile_name}", "wb") as zip:
                zip.write(zipdata)
        else:
            with open(f"{zip_path}\\trans_error-{ledger_id}.txt", "w") as errortxt:
                errortxt.write(f"Error : Not pid on translation for {ledger_id}")
            zipfile_name = f"{zip_path}\\trans_error-{ledger_id}.txt"
    with open(f"{data_path}\\tra_{ledger_id}{Year}{Month}.cfg","a+") as tra_file:
        tra_file.write(f"{ledger_id}{Year}{Month}\n")
    return f"{zipfile_name}"


def get_transfer_payload(ledger, period):
    logger.info("Generating transfer payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"BalanceTransferCrossLedger",
"ESSParameters":"{ledger},300000001414426,300000001414454,PTD,{period},{period},Y,N,Y,#NULL,#NULL,1",
"ReqstId":null
}}
'''




def tranferledgers(soap):
    logger.info("Transfer ledger call")
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/erpintegrations"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    return requests.post(url, data=soap, auth=(user, passw), headers=headerss)  # headers = headers



def Get_convertion(type, from_currency, to_currency, last_day):
    logger.info("Getting convertion")
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/currencyRates?finder=CurrencyRatesFinder;fromCurrency={from_currency},toCurrency={to_currency},userConversionType={type},startDate={last_day},endDate={last_day}&onlyData=True"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    return requests.get(url, auth=(user, passw), headers=headerss)


def Get_Close_SOAP(ledger_id, period):
    logger.info("Creating close ledger payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"ClosePeriod",
"ESSParameters":"{ledger_id},300000001414463,{ledger_id},{period},N,C,101,Y",
"ReqstId":null
}} '''

def Get_Close_CL_SOAP(ledger_id, period):
    logger.info("Creating close CL payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"ClosePeriod",
"ESSParameters":"{ledger_id},300000001414453,{ledger_id},{period},N,C,101,Y",
"ReqstId":null
}} '''


def Get_Ledger_Status(ledger_id,period):
    logger.info("Get ledger status call")
    url= f"{oracle_url}/fscmRestApi/resources/11.13.18.05/accountingPeriodStatusLOV?q=PeriodNameId={period};ApplicationId=101;LedgerId={ledger_id}"
    headerss = {"Content-Type": "application/json",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    return requests.get(url, auth=(user, passw), headers=headerss)


def Open_SOAP(ledger_id, period):
    logger.info("Open ledger payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"OpenPeriod",
"ESSParameters":"{ledger_id},300000001414463,{ledger_id},3002,101,P,{period},300000001412001,3002X300000001412001,Y",
"ReqstId":null
}}'''


def Open_CL_SOAP(ledger_id, period):
    logger.info("Open CL payload")
    return f'''{{
"OperationName":"submitESSJobRequest",
"JobPackageName":"/oracle/apps/ess/financials/generalLedger/programs/common",
"JobDefName":"OpenPeriod",
"ESSParameters":"{ledger_id},300000001414453,{ledger_id},3002,101,P,{period},300000001412001,3002X300000001412001,Y",
"ReqstId":null
}}'''



def zipdir(path, ziph):
    logger.info("Generating output zip")
    # ziph is zipfile handle
    for root, dirs, files in os.walk(path):
        for file in files:
            ziph.write(os.path.join(root, file), 
                       os.path.relpath(os.path.join(root, file), 
                                       os.path.join(path, '..')))


def clear_zip_dir():
    logger.info("Clearing directories")
    files_to_delete = os.listdir(zip_path)
    for item in files_to_delete:
        if item.endswith(".zip"):
            os.remove(os.path.join(zip_path, item))
    for item in files_to_delete:
        if item.endswith(".txt"):
            os.remove(os.path.join(zip_path, item))
    files_to_delete = os.listdir(ROOT_DIR)
    for item in files_to_delete:
        if item.endswith(".zip"):
            os.remove(os.path.join(ROOT_DIR, item))
    files_to_delete = os.listdir(zip_path)
    for item in files_to_delete:
        if item.endswith(".log"):
            os.remove(os.path.join(zip_path, item))
    files_to_delete = os.listdir(zip_path)
    for item in files_to_delete:
        if item.endswith(".xml"):
            os.remove(os.path.join(zip_path, item))
    files_to_delete = os.listdir(unzip_path)
    for item in files_to_delete:
        if item.endswith(".log"):
            os.remove(os.path.join(unzip_path, item))
    files_to_delete = os.listdir(unzip_path)
    for item in files_to_delete:
        if item.endswith(".xml"):
            os.remove(os.path.join(unzip_path, item))
    for item in files_to_delete:
        if item.endswith(".xlsx"):
            os.remove(os.path.join(unzip_path, item))
    for item in files_to_delete:
        if item.endswith(".txt"):
            os.remove(os.path.join(unzip_path, item))
 


        
def create_zip():
    logger.info("Creating zip")
    zip_file = zipfile.ZipFile(f'{ROOT_DIR}\\Output.zip', 'w', zipfile.ZIP_DEFLATED)
    zipdir(zip_path, zip_file)
    zip_file.close()
    return f'{ROOT_DIR}\\Output.zip'


def verify(data):
    logger.info(f"Trying to verify data {data}")
    data=data.upper()
    try:      
        Period, region = data.split(",")
        region = region.upper()
        Month, Year = Period.split("-")
        
        if region not in REGIONS:
            #print("Not in region")
            logger.info(f"Is not a valid region {region}")
            logger.info(f"Trying to push the log into the s3_bucket")
            Push_To_S3(f"{logs_path}\\log.log", "Process3", "Log")
            return f"Python Error : Invalid region"
        else:
            return f"{Period},{region}"
    except:
        logger.info(f"Error on subject Region or period not well defined")
        logger.info(f"Trying to push the log into the s3_bucket")
        Push_To_S3(f"{logs_path}\\log.log", "Process6", "Log")
        return f"Python Error : Region, Period or Ledger not well defined."



def remover(file):
    logger.info("Removing file")
    try:
        os.remove(f"{file}")
    except:
        pass
    
def Post_Journals():
    logger.info("POsting journal")
    rest_post = Create_AutoPost(Autopost_id)
    response_autopost = GL_Request(rest_post)
    Autopost_Reqid = json.loads(response_autopost.text)["ReqstId"]
    while True:
        Status = json.loads(Get_Status(Autopost_Reqid).text)["items"][0]["RequestStatus"]
        if Status == "SUCCEEDED":
            return "<li>Journals Posted</li>"
        if Status == "ERROR" or Status == "WARNING":
            logger.info("Error on autopost")
            return f"Python Error: Warning/Error on Autopost."



def remove_error():
    logger.info("Removing error file")
    try:
        os.remove(f"{data_path}\\error.txt")
    except:
        pass
    

def Save_Sheet(Dic, File, Sheet, Col, Row):
    logger.info("Saving xlsx file")
    df = pd.DataFrame.from_dict(Dic).T
    try:
        book = load_workbook(File)
    except:
        book = Workbook()
        book.save(File)
        book = load_workbook(File)
    writer = pd.ExcelWriter(File, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=Sheet, startcol=Col, startrow=Row)
    writer.save()



def get_mpa_report(file,filename):
        logger.info("Creating MPA report")
    #try:
        with zipfile.ZipFile(f"{file}", 'r') as zip_ref:
            zip_ref.extractall(f"{unzip_path}")
        pid = file.split("\\")[-1]
        pid = pid.split("_")[-1]
        pid = pid.split(".")[0]
        xml_file=f"{unzip_path}\\ESS_O_{pid}_BIP.xml"
        Xml_Data = open(xml_file,"r").read()
        dictionary = xmltodict.parse(Xml_Data)
        data = json.loads(json.dumps(dictionary))["XLAAPRPT"]
        String=""
        for i in data:
            print(f"\n\n\n{String}\n\n\n")
            String += get_string(i,data,0)
        #Save_Sheet(new_data,f"{unzip_path}\\MPA_RRPORT_TEST.xlsx",f"{Test}", 0, 0)
        print(String)
        data_lines= String.splitlines()
        logger.info("MPA REPORT DONE")
        write_pdf(data_lines,f"{filename}")
        return f"{filename}"
    #except:
    #    logger.info("Not exceptions detected")
    #    return ""


def get_mpa_report_d(file):
    #logger.info("Creating MPA report")
    #try:
        xml_file=f"{unzip_path}\\{file}"
        Xml_Data = open(xml_file,"r").read()
        dictionary = xmltodict.parse(Xml_Data)
        data = json.loads(json.dumps(dictionary))["XLAAPRPT"]
        String=""
        for i in data:
            print(f"\n\n\n{String}\n\n\n")
            String += get_string(i,data,0)
        #Save_Sheet(new_data,f"{unzip_path}\\MPA_RRPORT_TEST.xlsx",f"{Test}", 0, 0)
        print(String)
        data_lines= String.splitlines()
        logger.info("MPA REPORT DONE")
        write_pdf(data_lines,xml_file)
        return f"{unzip_path}\\mpa_report_test.xlsx"
    #except:
    #    logger.info("MPA cannot be created")
    #    return "OK"

def get_string(i,data,level):
    t=type(data[i])
    #print(f"I:{i},data[i]:{data[i]},Level:{level},Type:{t}")
    if type(data[i]) is str:
        spaces=""
        for l in range(level):
            spaces+="\t"
        value=data[i] 
        if value != None and value != "":     
            String = f"{spaces}{i} : {value}\n"
        else:
            return ""
        return String
    elif type(data[i]) is dict:
        #String=f"{i}\n\t"
        String=""
        for j in data[i]:
            String += get_string(j,data[i],level+1)
    else:
        #String = f"{i}\n"
        String=""
    return String



def write_pdf(data_lines,name):
    pdf = FPDF()
    pdf.add_page()
    for x in data_lines:
        if "NAME" in x or "LEDGER" in x or "WARNING" in x or "ERROR" in x:  
            pdf.set_font("Arial","BUI", size = 10)
        else:
            pdf.set_font("Arial", size = 10)
        sp = "\t" * x.count("\t")
        print(x.count("\t"),sp)
        x = f'\n{sp}'.join(x[i:i+100] for i in range(0, len(x), 100))
        print(x)
        y = x.splitlines()
        for x in y:
            pdf.cell(200, 10, txt = x, ln = 2, align = 'J')
    pdf.output(name) 
    
def create_report(CPE_zipfile):
    logger.info("Creating cpe report")
    try:
        with zipfile.ZipFile(f"{CPE_zipfile}", 'r') as zip_ref:
            zip_ref.extractall(f"{unzip_path}")
        pid = CPE_zipfile.split("\\")[-1]
        pid = pid.split("_")[-1]
        pid = pid.split(".")[0]
        xml_file=f"{unzip_path}\\ESS_O_{pid}_BIP.xml"
        Xml_Data = open(xml_file,"r").read()
        dictionary = xmltodict.parse(Xml_Data)
        data = json.loads(json.dumps(dictionary))["XLAPEXRPT"]["LEDGER_S"]["PERIOD_S"]["SOURCE_S"] 
        new_data={}
        count=0
        for i in range(len(data)):
            if type(data[i]['CLASS_S']) is dict:
                new_data[count]= data[i]['CLASS_S']
                count+=1
            elif type(data[i]['CLASS_S']) is list:
                for j in range(len(data[i]['CLASS_S'])):
                    new_data[count]= data[i]['CLASS_S'][j]
                    count+=1
            Save_Sheet(new_data,f"{unzip_path}\\Close_Exception_Report.xlsx",f"{count}", 0, 0)
        logger.info("Exceptions detected")
        return f"{unzip_path}\\Close_Exception_Report.xlsx"
    except:
        logger.info("Not exceptions detected")
        return "No Exceptions found"
    
        


def verify_daily(data):
    try:
        logger.info(f"Trying to verify data {data}")
        data=data.upper()
        Region = data.split(",")[1]
        if Region not in REGIONS:
            return f"Error: Region {Region} not defined:"             
    except:
        logger.info(f"Data received not well defined")
        logger.info(f"Trying to push the log into the s3_bucket")
        Push_To_S3(f"{logs_path}\\functions.log", "Process6", "Log")
        return "Python Error : Data Malformed."
    else:
        return data

def get_all_ledgers():
    logger.info("Reading Ledgers Id from Oracle")
    try:
        ledgers= {}
        ledgers_info = get_ledgers_info()
        #print(ledgers_info)
        #print(ledgers_info.text)
        ledger_json=json.loads(ledgers_info.text)
        ledger_json = ledger_json["items"]
        with open(f"{data_path}\\ledgers.txt","w") as ldgr_info:
            for i in ledger_json:
                info=f"{i}\n"
                ldgr_info.write(info)
    except:
        return "Error on set the entities // // "
    
    
    


def Get_L_Status(Period,ledger_id,name):
    logger.info(f"Getting Ledgers Status")
    #entities = get_entities("subledger") 
    ledger_status = verify_closed_periods(Period,ledger_id)
    ledger_json=json.loads(ledger_status.text)
    ledger_json = ledger_json["items"][0]
    #ledger_name =entities[ledger_id]["Name"]
    status = ledger_json["ClosingStatus"]
    if status == "C":
        status = "Closed"
    if status == "O":
        status = "Opened"
    if status == "N":
        status = "Never Opened"
    #name = entities[ledger_id]["Name"]
    return f"Status for the  ledger {name} on the period {Period} : {status}"



def verify_closed_periods(Period,Ledger_id):
    logger.info(f"Verifing Ledgers Status")
    url = f"{oracle_url}/fscmRestApi/resources/11.13.18.05/accountingPeriodStatusLOV?q=PeriodNameId={Period};ApplicationId=101;LedgerId={Ledger_id}"
    print(url)
    user,passw  = read_credentials()
    headerss = {"Content-Type": "application/json",
            "Connection": "Keep-Alive"
            }
    return requests.get(url, auth=(user, passw), headers=headerss)







def get_spreedsheet(soap):
    # print(soap)
    logger.info("Getting xlsx file to get the revalue id's")
    url = f"{oracle_url}/xmlpserver/services/ExternalReportWSSService?WSDL"
    headerss = {"Content-Type": "application/soap+xml",
                "Accept-Encoding": "gzip, deflate",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    resp = requests.post(url, data=soap, auth=(user, passw), headers=headerss)
    respcod = resp.text.split("Bytes>")[1].split("</ns2")[0]
    respdec = base64.b64decode(respcod)
    return respdec

'''

def get_report(report_type):
    if report_type == "Allocation":
        soap_report=Create_allocation_report_SOAP()
        xlsx_name = f"{config_path}\\allocation.xlsx"
    elif report_type == "Revalue":
        soap_report=Create_revalue_report_SOAP()
        xlsx_name = f"{config_path}\\revalue.xlsx"
    else: 
        return "No report type detected" 
    data = get_spreedsheet(soap_report)
    with open(f"{xlsx_name}", "wb") as xl:
        xl.write(data)
    read_file = pd.read_excel (f"{xlsx_name}")
    csv_file = xlsx_name.split("xlsx")[0]
    csv_file = f"{csv_file}csv"
    read_file.to_csv (f'{csv_file}', index = None, header=False)
    return f"{report_type} report created"
'''




def get_report(report_type):
    logger.info(f"Creating {report_type} report")
    if report_type == "Allocation":
        soap_report=Create_allocation_report_SOAP()
        xlsx_name = f"{config_path}\\allocation.csv"
    elif report_type == "Revalue":
        soap_report=Create_revalue_report_SOAP()
        xlsx_name = f"{config_path}\\revalue.csv"
    else: 
        return "No report type detected" 
    data = get_data(soap_report)
    with open(f"{xlsx_name}", "wb") as xl:
        data.replace(b'"',b"")
        xl.write(data)
    #read_file = pd.read_excel (f"{xlsx_name}")
    #csv_file = xlsx_name.split("xlsx")[0]
    #csv_file = f"{csv_file}csv"
    #read_file.to_csv (f'{csv_file}', index = None, header=False)
    return f"{report_type} report created"
        
        
     
        
        
def Create_allocation_report_SOAP():
    logger.info("Creating allocation SOAP")
    return f'''<soap:Envelope xmlns:pub="http://xmlns.oracle.com/oxp/service/PublicReportService" xmlns:soap="http://www.w3.org/2003/05/soap-envelope">
    <soap:Header></soap:Header>
    <soap:Body>
        <pub:runReport>
            <pub:reportRequest>
                <pub:parameterNameValues>
                    <pub:item>
                        <pub:name/>
                        <pub:values>
                            <pub:item/>
                        </pub:values>
                    </pub:item>
                </pub:parameterNameValues>
            <pub:reportAbsolutePath>/Custom/RPA/GW Allocation Rues/GW Allocation Rules.xdo</pub:reportAbsolutePath>
            <pub:sizeOfDataChunkDownload>-1</pub:sizeOfDataChunkDownload>
            </pub:reportRequest>
        </pub:runReport>
    </soap:Body>
</soap:Envelope>
'''

     
        
def Create_revalue_report_SOAP():
    logger.info("Creating revalue SOAP")
    user,passw = read_credentials()
    return f'''<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:v2="http://xmlns.oracle.com/oxp/service/v2">
   <soapenv:Header/>
   <soapenv:Body>
      <v2:runReport>
         <v2:reportRequest>
            <v2:attributeFormat>csv</v2:attributeFormat>
            <v2:byPassCache>y</v2:byPassCache>
            <v2:flattenXML>y</v2:flattenXML>
            <v2:reportAbsolutePath>/Custom/RPA/GW Revaluation Rules/GW Revaluation Rules.xdo</v2:reportAbsolutePath>
            <v2:sizeOfDataChunkDownload>-1</v2:sizeOfDataChunkDownload>
         </v2:reportRequest>
         <v2:userID>{user}</v2:userID>
         <v2:password>{passw}</v2:password>
      </v2:runReport>
   </soapenv:Body>
</soapenv:Envelope>
'''

 


def get_rev():
    logger.info("Getting revalue´s id")
    report = get_report("Revalue")
    if "No" not in report:
        csv_file = f"{config_path}\\revalue.csv"
        with open(csv_file,"r") as csv_data:
            revids = csv_data.read().splitlines()
        del revids[0]
        remove = ["SL", "Switzerland"  ,"Bermuda","Hong"]
        #data = any(element not in remove for element in data )
        revids = [element for element in revids if element not in remove ]
        revaluation={}
        with open(f"{config_path}\\revid.txt","r") as rev_data:
            revent = rev_data.read().splitlines()
            for i in revent:
                current_ent = i.split(",")[0]
                rev_name=i.split(",")[1]
                for j in revids:
                    if rev_name in j:
                        current_id = j.split(",")[0]
                        revaluation[current_ent] = {"Ledger": current_id, "Name": rev_name}
        return revaluation
                    
                    
                    
def get_data(soap):
    # print(soap)
    logger.info("Getting xlsx file to get the revalue id's")
    url = f"{oracle_url}:443/xmlpserver/services/v2/ReportService?wsdl"
    headerss = {"Content-Type": "application/soap+xml",
                "Accept-Encoding": "text, xml",
                "Connection": "Keep-Alive"
                }
    user,passw = read_credentials()
    resp = requests.post(url, data=soap, headers=headerss)
    print(resp.text)
    respcod = resp.text.split("Bytes>")[1].split("</report")[0]
    print(respcod)
    respdec = base64.b64decode(respcod)
    return respdec