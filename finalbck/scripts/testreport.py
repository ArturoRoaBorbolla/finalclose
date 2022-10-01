import finalclosefunctions
import logzero
import json
from logzero import logger
import os
import zipfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import xmltodict

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
zip_path = os.path.join(ROOT_DIR, "zip")
os.makedirs(zip_path, exist_ok=True)
config_path = os.path.join(ROOT_DIR, "Config")
os.makedirs(config_path, exist_ok=True)
logs_path = os.path.join(ROOT_DIR, "Logs")
os.makedirs(logs_path, exist_ok=True)
data_path = os.path.join(ROOT_DIR, "Data")
os.makedirs(data_path, exist_ok=True)
valid_response_Posted = "<ROWSET>\n<ROW>\n<P_LEDGER_ID>300000001414016</P_LEDGER_ID>\n<P_PERIOD_NAME>Aug-22</P_PERIOD_NAME>\n</ROW>\n</ROWSET>"
valid_length = len(valid_response_Posted.splitlines())
Fail_Flag_Posted = 0
unzip_path = os.path.join(ROOT_DIR, "Unzip")
os.makedirs(unzip_path, exist_ok=True)
general_data_path = os.path.join(ROOT_DIR, "..\\generaldata")
os.makedirs(general_data_path, exist_ok=True)
logzero.loglevel(logzero.INFO)
logzero.logfile(f"{logs_path}\\log.log")
logzero.json()

MJE_test_ID="300000164285311"

def Save_Sheet(Dic, File, Sheet, Col, Row):
    logger.info("Saving xlsx file")
    df = pd.DataFrame.from_dict(Dic).T
    try:
        book = load_workbook(File)
    except:
        book = Workbook()
        book.save(File)
        book = load_workbook(File)
    writer = pd.ExcelWriter(File, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name=Sheet, startcol=Col, startrow=Row)
    writer.save()

def create_report():
    try:
        xml_file=f"{ROOT_DIR}\\scripts\\ESS_O_7100435_BIP.xml"
        Xml_Data = open(xml_file,"r").read()
        dictionary = xmltodict.parse(Xml_Data)
        new_data={}
    except:
        return "Cant open XML file"
    else:
        try:
            data = json.loads(json.dumps(dictionary))["XLAPEXRPT"]["LEDGER_S"]["PERIOD_S"]["SOURCE_S"] 
            count=0
            for i in range(len(data)):
                if type(data[i]['CLASS_S']) is dict:
                    new_data[count]= data[i]['CLASS_S']
                    count+=1
                elif type(data[i]['CLASS_S']) is list:
                    for j in range(len(data[i]['CLASS_S'])):
                        new_data[count]= data[i]['CLASS_S'][j]
                        count+=1
                Save_Sheet(new_data,f"{unzip_path}\\Close_Exception_Report.xlsx",f"{count}", 0, 0)
            logger.info("Exceptions detected")
            return f"{unzip_path}\\Close_Exception_Report.xlsx"
        except:
            data = json.loads(json.dumps(dictionary))["XLAPEXRPT"]
            for i in data.keys():
                if data[i] != None:
                    new_data[i]= data[i]
                if i =="G_1" or i =="G_2":
                    for j in data[i].keys():
                        new_data[j]= data[i][j]
            print(new_data)
            Save_Sheet(new_data,f"{unzip_path}\\No_Close_Exception_Report.xlsx",f"No Exception", 0, 0)
            
             
        
    